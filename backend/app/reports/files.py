import csv
import io
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Iterable

from fastapi import UploadFile


PROJECT_ROOT = Path(__file__).resolve().parents[3]
UPLOAD_ROOT = PROJECT_ROOT / "uploads"
SUBMISSION_DIR = UPLOAD_ROOT / "submissions"
TRAINING_DIR = UPLOAD_ROOT / "ai_training"
TEMPLATE_DIR = UPLOAD_ROOT / "report_templates"


for directory in (SUBMISSION_DIR, TRAINING_DIR, TEMPLATE_DIR):
    directory.mkdir(parents=True, exist_ok=True)


def timestamp_name(filename: str) -> str:
    safe = Path(filename or "upload.bin").name.replace("/", "_").replace("\\", "_")
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    return f"{stamp}_{safe}"


async def store_upload(upload: UploadFile, folder: Path) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / timestamp_name(upload.filename or "upload.bin")
    with path.open("wb") as target:
        shutil.copyfileobj(upload.file, target)
    await upload.close()
    return path


def relative_path(path: Path | None) -> str | None:
    if not path:
        return None
    try:
        return str(path.relative_to(PROJECT_ROOT))
    except ValueError:
        return str(path)


def resolve_stored_path(value: str | None) -> Path | None:
    if not value:
        return None
    path = Path(value)
    if not path.is_absolute():
        path = PROJECT_ROOT / path
    return path


def extract_text_from_path(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".csv", ".json"}:
            return path.read_text(encoding="utf-8", errors="ignore")
        if suffix == ".pdf":
            return extract_pdf(path)
        if suffix in {".docx", ".doc"}:
            return extract_docx(path)
        if suffix in {".xlsx", ".xls"}:
            return extract_xlsx(path)
    except Exception as exc:
        return f"[Text extraction failed for {path.name}: {exc}]"
    return f"[Uploaded file stored as {path.name}. Text extraction is not available for this file type.]"


def extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        try:
            from PyPDF2 import PdfReader
        except Exception:
            return "[PDF text extraction requires pypdf or PyPDF2.]"

    reader = PdfReader(str(path))
    pages = []
    for page in reader.pages[:80]:
        pages.append(page.extract_text() or "")
    return "\n".join(pages).strip()


def extract_docx(path: Path) -> str:
    try:
        from docx import Document as DocxDocument
    except Exception:
        return "[DOCX text extraction requires python-docx.]"
    doc = DocxDocument(str(path))
    return "\n".join(paragraph.text for paragraph in doc.paragraphs).strip()


def extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return "[Excel text extraction requires openpyxl.]"
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    rows = []
    for sheet in workbook.worksheets[:8]:
        rows.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(max_row=250, values_only=True):
            values = [str(cell) for cell in row if cell is not None]
            if values:
                rows.append(" | ".join(values))
    return "\n".join(rows).strip()


def compact_text(text: str, limit: int = 14000) -> str:
    cleaned = " ".join((text or "").split())
    return cleaned[:limit]


def dumps_ids(values: Iterable[int]) -> str:
    return json.dumps([int(value) for value in values])


def loads_ids(value: str | None) -> list[int]:
    if not value:
        return []
    try:
        return [int(item) for item in json.loads(value)]
    except Exception:
        return []
